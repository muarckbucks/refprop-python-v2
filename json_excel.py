import pandas as pd
import json
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Alignment

fichero_json = r"resultados\resultados_ciclo_basico.json"
fichero_excel = r"resultados\resultados_ciclo_basico.xlsx"

# Cargar json
with open(fichero_json, "r", encoding="utf-8") as f:
    data = json.load(f)

# Crear ExcelWriter
keys_filtradas = ["COP", "VCC", "error"]
key_composicion = "mezcla"

with pd.ExcelWriter(fichero_excel, engine="openpyxl") as writer:
    for hoja, columnas in data.items():
        columnas_expandidas = []
        max_filas = 0

        # Preparar datos para cada columna
        for col_name, bloques in columnas.items():
            filas = []
            for bloque in bloques:
                for key in keys_filtradas:
                    if key in bloque:
                        if isinstance(bloque[key], float):
                            filas.append([key, round(bloque[key], 3)])
                        else:
                            filas.append([key, bloque[key]])
                filas.append(["", ""])  # espacio entre bloques
            columnas_expandidas.append((col_name, filas))
            max_filas = max(max_filas, len(filas))

        # Rellenar con filas vacías
        for i in range(len(columnas_expandidas)):
            col_name, filas = columnas_expandidas[i]
            filas.extend([["", ""]] * (max_filas - len(filas)))
            columnas_expandidas[i] = (col_name, filas)

        # Construir DataFrame
        df_dict = {}
        # Columna extra para "Composición" (temporal, rellena con "")
        df_dict["Composición"] = [""] * max_filas
        for col_name, filas in columnas_expandidas:
            for i, subcol in enumerate(["key", "valor"]):
                df_dict[f"{col_name}_{subcol}"] = [fila[i] for fila in filas]

        df = pd.DataFrame(df_dict)
        df.to_excel(writer, sheet_name=hoja, index=False)

# Formato con openpyxl hoja por hoja
wb = load_workbook(fichero_excel)
for hoja in wb.sheetnames:
    ws = wb[hoja]

    # Reconstruir columnas_expandidas y bloques
    columnas = data[hoja]
    columnas_expandidas = []
    bloques_info = []  # para composición (fila inicio, fila fin, valor)
    max_filas = 0
    for col_name, bloques in columnas.items():
        filas = []
        bloque_inicio = 2  # empieza en fila 2 porque fila 1 es header
        for bloque in bloques:
            bloque_filas = []
            for key in keys_filtradas:
                if key in bloque:
                    if isinstance(bloque[key], float):
                        bloque_filas.append([key, round(bloque[key], 3)])
                    else:
                        bloque_filas.append([key, bloque[key]])
            filas.extend(bloque_filas)
            filas.append(["", ""])  # espacio entre bloques

            # Guardar info del bloque para composición
            if key_composicion in bloque:
                # Sacar string de composición list[float] -> str
                string_comp = ""
                for comp in bloque[key_composicion]:
                    string_comp += f"{(comp*100):.0f}% "

                filas_bloque = len(bloque_filas)
                bloques_info.append((bloque_inicio, bloque_inicio + filas_bloque - 1, string_comp[:-1]))
            bloque_inicio += len(bloque_filas) + 1

        columnas_expandidas.append((col_name, filas))
        max_filas = max(max_filas, len(filas))

    # Combinar celdas del header y renombrar columnas
    col_index = 2  # columna 1 es "Composición"
    for col_name, _ in columnas_expandidas:
        start = col_index
        end = col_index + 1
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        ws.cell(row=1, column=start).value = col_name
        col_index += 2

    # Poner título "Composición" en la primera fila, centrado
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1)
    ws.cell(row=1, column=1).value = "Composición"

    # Añadir los valores de composición combinando verticalmente cada bloque
    for start_row, end_row, valor in bloques_info:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        ws.cell(row=start_row, column=1).value = valor

    # Centrar todas las celdas
    alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = alignment

    # Ajustar ancho columnas
    for i, column_cells in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = 12

    # Inmovilizar primar fila y columna
    ws.freeze_panes = "B2"

wb.save(fichero_excel)