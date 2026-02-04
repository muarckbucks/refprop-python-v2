from refprop_utils import * 
from typing import Any, Iterable
from pprint import pprint
import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font

def calcular_ciclo_basico(
    fluido: str | list[str],
    mezcla: list[float],
    temperaturas_agua: dict[str, list[float]],
    approach_ini: float = 6.5, # Provar
    approach_max: float = 20,
    step: float = 0.5
) -> CicloOutput:

    approach = approach_ini

    while approach < approach_max:
        resultado = calcular_ciclo(fluido, mezcla, temperaturas_agua, approach)

        if resultado.error is not None:
            return resultado

        if resultado.pinch >= 1:
            return resultado

        approach += step

    return CicloOutput(fluido=resultado.fluido,
                       mezcla=resultado.mezcla,
                       temperaturas_agua=temperaturas_agua,
                       error="PinchBajo")

def calcular_ciclo(fluido: str | list[str], mezcla: list[float],
                   temperaturas_agua: dict[str, list[float]], approach_k: float) -> CicloOutput:
    
    resultado_basico: dict[str, Any] = {}

    resultado_basico["fluido"] = fluido
    resultado_basico["mezcla"] = mezcla
    resultado_basico["temperaturas_agua"] = temperaturas_agua

    [t_hw_in, t_hw_out] = temperaturas_agua["t_hw"]
    [t_cw_in, t_cw_out] = temperaturas_agua["t_cw"]

    ap_k = approach_k
    ap_0 = 3

    SH = 5
    SUB = 1

    t3 = t_hw_in + ap_k
    T_crit = rprop(fluido, "Tcrit", mezcla, T = 0, H = 0)

    try:
        if t3 > T_crit:
            raise ErrorTemperaturaTranscritica(f"Temperatura transcrítica en el punto de descarga: {t3:.1f}ºC > {T_crit:.1f}ºC")

        PK = rprop(fluido, "P", mezcla, T = t3 + SUB, Q = 0)

        # Punto 3
        P3 = TPoint(fluido, mezcla, T = t3, P = PK)
        P3.calcular("H", "D")


        # Punto 4
        P4 = TPoint(fluido, mezcla, H = P3.H, T = t_cw_out - ap_0)
        P4.calcular("P", "H", "T")
        P0 = P4.P

        # Rendimiento isentrópico
        rend_iso_h = 0.6

        # Punto 1
        t_sat_1 = rprop(fluido, "T", mezcla, P = P0, Q = 1)
        P1 = TPoint(fluido, mezcla, T = t_sat_1 + SH, P = P0)
        P1.calcular("H", "S", "V")

        # Punto 2
        h_2_s = rprop(fluido, "H", mezcla, P = PK, S = P1.S)
        h_2 = P1.H + (h_2_s - P1.H)/rend_iso_h
        P2 = TPoint(fluido, mezcla, P = PK, H = h_2)
        P2.calcular("H", "Q", "D", "T")
        if P2.Q <= 1:
            raise ErrorPuntoBifasico("El punto de descarga cae en la zona bifásica")

        # COP y VCC
        COP = (P2.H - P3.H)/(P2.H - P1.H)
        VCC = (P2.H - P1.H)/P1.V

        # Puntos saturados
        Pk_liq_sat = TPoint(fluido, mezcla, P = PK, Q = 0)
        Pk_vap_sat = TPoint(fluido, mezcla, P = PK, Q = 1)

        P0_liq_sat = TPoint(fluido, mezcla, P = P0, Q = 0)
        P0_vap_sat = TPoint(fluido, mezcla, P = P0, Q = 1)

        puntos_saturados = [Pk_liq_sat, Pk_vap_sat, P0_liq_sat, P0_vap_sat]
        # Caudales
        P_hw_in = TPoint("WATER", P = 1, T = t_hw_in)
        P_hw_out = TPoint("WATER", P = 1, T = t_hw_out)
        P_cw_in = TPoint("ETHYLENEGLYCOL", P = 1, T = t_cw_in)
        P_cw_out = TPoint("ETHYLENEGLYCOL", P = 1, T = t_cw_out)

        # Relaciones másicas
        ratio_m_GlycolHot_R = (P2.H - P3.H)/(P_hw_out.H - P_hw_in.H)
        ratio_m_GlycolCold_R = (P1.H - P4.H)/(P_cw_in.H - P_cw_out.H)

        # Relaciones volumétricas
        ratio_v_GlycolHot_R = ratio_m_GlycolHot_R * P2.D/P_hw_in.D
        ratio_v_GlycolCold_R = ratio_m_GlycolCold_R * P3.D/P_cw_in.D

        # Pinch
        h_water_pinch = P_hw_out.H - 1/ratio_m_GlycolHot_R * (P2.H - Pk_vap_sat.H)
        P_water_pinch = TPoint("WATER", P = 1, H = h_water_pinch)
        pinch = Pk_vap_sat.T - P_water_pinch.T

        # Glide
        glide_k = Pk_vap_sat.T - Pk_liq_sat.T
        glide_0 = P0_vap_sat.T - P4.T

        puntos = {
            "1": P1,
            "2": P2,
            "3": P3,
            "4": P4,
        }
        

        resultados_adicionales = {
            "COP": COP,
            "VCC": VCC,
            "puntos": puntos,
            "puntos_sat": puntos_saturados,
            "caudales_mas": [ratio_m_GlycolHot_R, ratio_m_GlycolCold_R],
            "caudales_vol": [ratio_v_GlycolHot_R, ratio_v_GlycolCold_R],
            "pinch": pinch,
            "glide": [glide_k, glide_0],
            "approach_k": ap_k,
            "error": None,
        }



        resultado = resultado_basico | resultados_adicionales

        output = CicloOutput(**resultado)

    except ErrorPuntoBifasico:
        resultado = resultado_basico | {"error": "Bifásico"}
        output = CicloOutput(**resultado)

    except ErrorTemperaturaTranscritica:
        resultado = resultado_basico | {"error": "Transcrítico"}
        output = CicloOutput(**resultado)
    
    except ZeroDivisionError:
        resultado = resultado_basico | {"error": "División 0"}
        output = CicloOutput(**resultado)

    except RuntimeError:
        resultado = resultado_basico | {"error": "REFPROP"}
        output = CicloOutput(**resultado)

    return output

def worker_calcular(args):
    # Check REFPROP handle in the refprop_utils module (initializer sets this per process)
    import refprop_utils
    if refprop_utils.RP is None:
        raise RuntimeError("REFPROP no inicializado en el worker")

    fluido, mezcla, temperaturas_agua = args
    res = calcular_ciclo_basico(fluido, mezcla, temperaturas_agua)
    return serializar(res)

def calcular_mezclas(fichero_json: str, posibles_refrigerantes: list[str], temperaturas_agua: dict[str, list[float]]):
    n_calcs = 41
    # Inicializar diccionario de resultados
    resultados: dict[str, dict[str, list[CicloOutput]]] = {}
    for ref_a in posibles_refrigerantes:
        resultados[ref_a] = {}
        for ref_b in posibles_refrigerantes:
            if ref_a != ref_b:
                resultados[ref_a][ref_b] = []
                [resultados[ref_a][ref_b].append(0) for _ in range(n_calcs)]

    # Calcular mezclas de refrigerantes
    for index_a, ref_a in enumerate(posibles_refrigerantes[:-1]):

        for ref_b in posibles_refrigerantes[index_a + 1:]:

            props_a = list(np.linspace(0, 1, n_calcs))
            props_a = [float(x) for x in props_a]
            
            for index_prop, prop_a in enumerate(props_a):
                prop_b = 1 - prop_a
                mezcla = [prop_a, prop_b]

                resultado = calcular_ciclo_basico([ref_a, ref_b], mezcla, temperaturas_agua)
                # Ir imprimiendo resultados
                string_comp = ""

                # Comprobar si no ha dado error el cálculo
                if resultado.error is None:
                    for fluid, comp in zip(resultado.fluido, resultado.mezcla):
                        string_comp += f"{fluid}: {(comp*100):.1f}%, "
                    print(string_comp + f"COP = {resultado.COP:.3f}")
                else:
                    for fluid, comp in zip(resultado.fluido, resultado.mezcla):
                        string_comp += f"{fluid}: {(comp*100):.1f}%, "
                    print(string_comp + f"ERROR = {resultado.error}")                    
                
                # Graficar resultado
                # if resultado.error is None:
                #     puntos_PH(list(resultado.puntos.values()), 1.5, 0.2)

                resultados[ref_a][ref_b][index_prop] = resultado
                resultados[ref_b][ref_a][n_calcs - 1 - index_prop] = resultado
 
    # Guardar resultados en json
    with open(fichero_json, "w", encoding="utf-8") as f:
        json.dump(serializar(resultados), f, ensure_ascii=False, indent=2)

def json_a_excel(fichero_json, fichero_excel):

    # Cargar json
    with open(fichero_json, "r", encoding="utf-8") as f:
        data: dict[str, dict[str, dict[str, dict[str, Any]]]] = json.load(f)

    # Crear ExcelWriter
    keys_filtradas = ["COP", "VCC", "error"]
    key_composicion = "mezcla"

    with pd.ExcelWriter(fichero_excel, engine="openpyxl") as writer:
        for hoja, columnas in data.items():
            columnas_expandidas = []
            max_filas = 0

            # Preparar datos para cada columna
            for col_name, bloques in columnas.items():
                filas = []
                for bloque in bloques:
                    bloque = bloque["__data__"]
                    for key in keys_filtradas:
                        if key in bloque:
                            if isinstance(bloque[key], float):
                                filas.append([key, round(bloque[key], 3)])
                            else:
                                filas.append([key, bloque[key]])
                    filas.append(["", ""])  # espacio entre bloques
                columnas_expandidas.append((col_name, filas))
                max_filas = max(max_filas, len(filas))

            # Rellenar con filas vacías
            for i in range(len(columnas_expandidas)):
                col_name, filas = columnas_expandidas[i]
                filas.extend([["", ""]] * (max_filas - len(filas)))
                columnas_expandidas[i] = (col_name, filas)

            # Construir DataFrame
            df_dict = {}
            # Columna extra para "Composición" (temporal, rellena con "")
            df_dict["Composición"] = [""] * max_filas
            for col_name, filas in columnas_expandidas:
                for i, subcol in enumerate(["key", "valor"]):
                    df_dict[f"{col_name}_{subcol}"] = [fila[i] for fila in filas]

            df = pd.DataFrame(df_dict)
            df.to_excel(writer, sheet_name=hoja, index=False)

    # Formato con openpyxl hoja por hoja
    wb = load_workbook(fichero_excel)
    for hoja in wb.sheetnames:
        ws = wb[hoja]

        # Reconstruir columnas_expandidas y bloques
        columnas = data[hoja]
        columnas_expandidas = []
        bloques_info = []  # para composición (fila inicio, fila fin, valor)
        max_filas = 0
        for col_name, bloques in columnas.items():
            filas = []
            bloque_inicio = 2  # empieza en fila 2 porque fila 1 es header
            for bloque in bloques:
                bloque = bloque["__data__"]
                bloque_filas = []
                for key in keys_filtradas:
                    if key in bloque:
                        if isinstance(bloque[key], float):
                            bloque_filas.append([key, round(bloque[key], 3)])
                        else:
                            bloque_filas.append([key, bloque[key]])
                filas.extend(bloque_filas)
                filas.append(["", ""])  # espacio entre bloques

                # Guardar info del bloque para composición
                if key_composicion in bloque:
                    # Sacar string de composición list[float] -> str
                    string_comp = ""
                    for comp in bloque[key_composicion]:
                        string_comp += f"{(comp*100):.0f}% "

                    filas_bloque = len(bloque_filas)
                    bloques_info.append((bloque_inicio, bloque_inicio + filas_bloque - 1, string_comp[:-1]))
                bloque_inicio += len(bloque_filas) + 1

            columnas_expandidas.append((col_name, filas))
            max_filas = max(max_filas, len(filas))

        # Combinar celdas del header y renombrar columnas
        col_index = 2  # columna 1 es "Composición"
        for col_name, _ in columnas_expandidas:
            start = col_index
            end = col_index + 1
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
            ws.cell(row=1, column=start).value = col_name
            col_index += 2

        # Poner título "Composición" en la primera fila, centrado
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1)
        ws.cell(row=1, column=1).value = "Composición"

        # Añadir los valores de composición combinando verticalmente cada bloque
        for start_row, end_row, valor in bloques_info:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            ws.cell(row=start_row, column=1).value = valor

        # Centrar todas las celdas
        alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = alignment

        # Ajustar ancho columnas
        for i, column_cells in enumerate(ws.columns, 1):
            ws.column_dimensions[get_column_letter(i)].width = 12

        # Inmovilizar primar fila y columna
        ws.freeze_panes = "B2"

    wb.save(fichero_excel)

def refinar_mezclas(fichero_json, fichero_json_fino, fichero_txt):

    with open(fichero_json, "r", encoding="utf-8") as f:
        data_ini: dict[str, dict[str, list[dict[str, Any]]]] = json.load(f)

    data: dict[str, dict[str, list[CicloOutput]]] = deserializar(data_ini)

    temperaturas_agua = list(list(data.values())[0].values())[0][0].temperaturas_agua

    posibles_refrigerantes = list(data.keys())

    # Inicializar diccionario de resultados
    resultado_fino: dict[str, dict[str, CicloOutput]] = {}
    for ref_a in posibles_refrigerantes:
        resultado_fino[ref_a] = {}


    refrigerantes_revisados: set[str] = set()

    # Calcular VCC de referencia
    margen_vcc = 0.3
    vcc_propano = calcular_ciclo_basico("PROPANE", [1.0],
                                        temperaturas_agua).VCC
    vcc_min = (1 - margen_vcc) * vcc_propano
    vcc_max = (1 + margen_vcc) * vcc_propano
    # Calcular COP de propano
    cop_propano = calcular_ciclo_basico("PROPANE", [1.0],
                                        temperaturas_agua).COP

    salto = 0.025

    for ref_1, sub_data_1 in data.items():

        refrigerantes_revisados.add(ref_1)

        for ref_2, lista_resultados in sub_data_1.items():
            
            if ref_2 not in refrigerantes_revisados:

                fluido = lista_resultados[0].fluido
                # Aplicar filtros
                lista = lista_resultados
                # Quitar errores
                lista = [res for res in lista if res.error is None]
                # VCC
                if lista:
                    lista = [res for res in lista if (vcc_min <= res.VCC <= vcc_max)]
                # Temperatura descarga
                if lista:
                    lista = [res for res in lista if res.puntos["2"].T < 130]
                # Pinch > 0
                if lista:
                    lista = [res for res in lista if res.pinch > 0]
                # Glide < 10ºC
                lista = [res for res in lista if res.glide[0] < 10 and res.glide[1] < 10]
                # Guardar 2 resultados con mayor COP
                lista_resultados = sorted(lista, key = lambda resultado: resultado.COP, reverse = True)[:2]

                comps = []
                # Comprobar cuántos resultados dan COP correcto
                if len(lista_resultados) == 2: # Si hay 2 elementos coger los que tienen más COP
                    

                    # Pero si la diferencia es muy grande buscar entre los dos puntos
                    if abs(lista_resultados[0].mezcla[0] - lista_resultados[1].mezcla[0]) <= 0.10:

                        [comp_1, comp_2] = [res.mezcla for res in lista_resultados]
                        comps.append([comp_1, comp_2])
                    
                    else:
                        # Crear composiciones de +-5% alrededor de cada punto
                        comp_1 = [max(lista_resultados[0].mezcla[0] - 0.05, 0), 1 - max(lista_resultados[0].mezcla[0] - 0.05, 0)]
                        comp_2 = [min(lista_resultados[0].mezcla[0] + 0.05, 1), 1 - min(lista_resultados[0].mezcla[0] + 0.05, 1)]
                        comp_3 = [max(lista_resultados[1].mezcla[0] - 0.05, 0), 1 - max(lista_resultados[1].mezcla[0] - 0.05, 0)]
                        comp_4 = [min(lista_resultados[1].mezcla[0] + 0.05, 1), 1 - min(lista_resultados[1].mezcla[0] + 0.05, 1)]
                        
                        comps.append([comp_1, comp_2])
                        comps.append([comp_3, comp_4])

                elif len(lista_resultados) == 1:
                    # Si solo hay un elemento ir desde +- 5% composición
                    comp_1 = [max(lista_resultados[0].mezcla[0] - 0.05, 0), 1 - max(lista_resultados[0].mezcla[0] - 0.05, 0)]
                    comp_2 = [min(lista_resultados[0].mezcla[0] + 0.05, 1), 1 - min(lista_resultados[0].mezcla[0] + 0.05, 1)]
                    comps.append([comp_1, comp_2])

                else: # len(lista_resultados) == 0
                    # Si no hay resultados saltar la mezcla
                    continue
            
                # Iterar para cada rango de composiciones (solo útil cuando 
                # las 2 soluciones con más COP están alejadas y hay que 
                # crear 2 rangos diferentes)
                for comp_i in comps:
                    # Crear rango de composiciones que vaya desde comp1 a comp2
                    [comp_1, comp_2] = comp_i
                    salto = 0.005
                    range_comp: list[float] = [min(comp_1[0], comp_2[0]) + x * salto for x in range(int(round(abs(comp_1[0] - comp_2[0])/salto)) + 1)]

                    resultados: list[CicloOutput] = []

                    for comp in range_comp:
                        mezcla = [comp, 1 - comp]
                        resultado = calcular_ciclo_basico(fluido, mezcla, temperaturas_agua) # BUG calculant el COP hi ha una divisió per 0 ../(PS.H - P1.H)
                        resultados.append(resultado)

                        string_comp = ""
                        if resultado.error is None:
                            for fluid, comp in zip(resultado.fluido, resultado.mezcla):
                                string_comp += f"{fluid}: {(comp*100):.1f}%, "
                            print(string_comp + f"COP = {resultado.COP:.3f}")
                        else:
                            for fluid, comp in zip(resultado.fluido, resultado.mezcla):
                                string_comp += f"{fluid}: {(comp*100):.1f}%, "
                            print(string_comp + f"ERROR = {resultado.error}")

                
                # Filtrar otra vez por errores, vcc, temp. descarga, pinch y glide
                resultados_temp_1 = [res for res in resultados if res.error is None]
                resultados_temp_2 = [res for res in resultados_temp_1 if (vcc_min <= res.VCC <= vcc_max)]
                resultados_temp_1 = [res for res in resultados_temp_2 if res.puntos["2"].T < 130]
                resultados_temp_2 = [res for res in resultados_temp_1 if res.pinch > 0]
                resultados_temp_1 = [res for res in resultados_temp_2 if res.glide[0] < 10 and res.glide[1] < 10]
                resultados_temp_2 = [res for res in resultados_temp_1 if res.puntos["2"].P < 25] 
                
                lista_mejores_resultados = []
                # Quedarse con el COP más grande
                if resultados_temp_2 != []:

                    res_mayor_COP: CicloOutput = max(resultados_temp_2, key=lambda r: r.COP)
                    
                    resultado_fino[ref_1][ref_2] = res_mayor_COP
                    resultado_fino[ref_2][ref_1] = res_mayor_COP

                    string_comp = ""
                    for fluid, comp in zip(res_mayor_COP.fluido, res_mayor_COP.mezcla):
                        string_comp += f"{fluid}: {(comp*100):.1f}%, "
                    print("\nProporción VÁLIDA con más COP: " + string_comp + f"COP = {res_mayor_COP.COP:.3f}")
                    if res_mayor_COP.COP > cop_propano:
                        print(f"COP {((res_mayor_COP.COP/cop_propano-1)*100):.1f}% más GRANDE que propano\n")
                    else:
                        print(f"COP {(-(res_mayor_COP.COP/cop_propano-1)*100):.1f}% más PEQUEÑO que propano\n")
                    lista_mejores_resultados.append(res_mayor_COP)

    # Guardar resultados en json
    with open(fichero_json_fino, "w", encoding = "utf-8") as f:
        json.dump(serializar(lista_mejores_resultados), f, ensure_ascii=False, indent=2)

    # Guardar resultado en txt
    string_res: list[str] = []

    for res in lista_mejores_resultados:
        string_comp = ""

        for fluid, comp in zip(res.fluido, res.mezcla):
            string_comp += f"{fluid}: {abs((comp*100)):.0f}%, "

        proporcion = (res.COP / cop_propano - 1) * 100

        if proporcion >= 0:
            string_comp[:-2] += f"\nCOP {proporcion:.2f}% más GRANDE que el propano\n\n"
        else:
            string_comp[:-2] += f"\nCOP {-proporcion:.2f}% más PEQUEÑO que el propano\n\n"

        string_res.append(string_comp)

    with open(fichero_txt, "w",encoding="utf-8") as f:
        f.writelines(string_res)

def json_a_excel_fino(
    fichero_json: str,
    fichero_excel: str,
    ancho_col_key: float = 20,
    ancho_col_value: float = 30,
    ancho_col_separador: float = 5,
) -> None:

    keys_filtradas = ["COP", "VCC", "pinch", "glide"]

    with open(fichero_json, "r", encoding="utf-8") as f:
        data: dict[str, dict[str, dict[str, Any]]] = json.load(f)

    wb = Workbook()
    ws = wb.active

    align_center = Alignment(horizontal="center", vertical="center")
    bold = Font(bold=True)

    col_actual = 1

    for titulo_grupo, bloques in data.items():
        # Anchos de columnas del grupo
        ws.column_dimensions[get_column_letter(col_actual)].width = ancho_col_key
        ws.column_dimensions[get_column_letter(col_actual + 1)].width = ancho_col_value

        # Título de grupo
        ws.merge_cells(
            start_row=1,
            start_column=col_actual,
            end_row=1,
            end_column=col_actual + 1
        )
        cell = ws.cell(row=1, column=col_actual, value=titulo_grupo)
        cell.alignment = align_center
        cell.font = bold

        fila = 2

        for titulo_bloque, pares in bloques.items():
            pares = pares["__data__"]
            if pares != {}:
                # Título del bloque
                ws.merge_cells(
                    start_row=fila,
                    start_column=col_actual,
                    end_row=fila,
                    end_column=col_actual + 1
                )
                cell = ws.cell(row=fila, column=col_actual, value=titulo_bloque)
                cell.alignment = align_center
                fila += 1

                fluido = pares["fluido"]
                mezcla = pares["mezcla"]

                str_mezcla = ""
                for prop, ref in zip(mezcla, fluido):
                    str_mezcla += f"{(prop*100):.1f}% {ref} + "
                str_mezcla = str_mezcla[:-3]

                # Contenido
                for k in pares:
                    if k not in keys_filtradas:
                        continue
                    c_key = ws.cell(row=fila, column=col_actual, value=k)
                    if k == "glide":
                        texto = ""
                        for v in pares[k]:
                            texto += f"{v:.2f}ºC "
                        c_val = ws.cell(row=fila, column=col_actual + 1, value= texto)
                    elif k == "pinch":
                        c_val = ws.cell(row=fila, column=col_actual + 1, value= f"{pares[k]:.2f} ºC")
                    else:
                        c_val = ws.cell(row=fila, column=col_actual + 1, value= round(pares[k], 3))
                    c_key.alignment = align_center
                    c_val.alignment = align_center
                    fila += 1
                
                c_key = ws.cell(row=fila, column=col_actual, value = "Composición")
                c_val = ws.cell(row=fila, column=col_actual + 1, value = str_mezcla)
                c_key.alignment = align_center
                c_val.alignment = align_center
                fila += 1

                fila += 1  # línea en blanco entre bloques

        # Columna separadora
        sep_col = col_actual + 2
        ws.column_dimensions[get_column_letter(sep_col)].width = ancho_col_separador

        col_actual += 3

    ws.freeze_panes = "A2"

    wb.save(fichero_excel)



def main():
    init_refprop()
    # DATOS BÁSICOS
    t_hw_in = 47
    t_hw_out = 55

    t_cw_in = 0
    t_cw_out = -3

    temperaturas_agua = {
        "t_hw": [t_hw_in, t_hw_out],
        "t_cw": [t_cw_in, t_cw_out]
    }

    posibles_refrigerantes = ["PROPANE", "BUTANE", "ISOBUTANE", "PROPYLENE", "DME"]


    fichero_json = r"resultados\res_ciclo_basico.json"
    fichero_json_fino = r"resultados\res_finos_basico.json"
    fichero_excel = r"resultados\res_ciclo_basico.xlsx"
    fichero_excel_fino = r"resultados\res_finos_basico.xlsx"
    fichero_txt = r"resultados\res_finos_basico.txt"

    # Llamar a las funciones
    calcular_mezclas(fichero_json, posibles_refrigerantes, temperaturas_agua)

    json_a_excel(fichero_json, fichero_excel)

    refinar_mezclas(fichero_json, fichero_json_fino)

    json_a_excel_fino(fichero_json_fino, fichero_excel_fino, fichero_txt)



if __name__ == "__main__":
    main()


