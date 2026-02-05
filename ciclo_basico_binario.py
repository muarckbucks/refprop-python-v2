from refprop_utils import * 
from typing import Any
import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
import matplotlib.pyplot as plt
from concurrent.futures import ProcessPoolExecutor
from tqdm import tqdm
from pprint import pprint

def calcular_ciclo_basico(
    fluido: str | list[str],
    mezcla: list[float],
    water_config: str,
    approach_ini: float = 6.5, # Provar
    approach_max: float = 20,
    step: float = 0.5
) -> CicloOutput:

    approach = approach_ini

    while approach < approach_max:
        resultado = calcular_ciclo(fluido, mezcla, water_config, approach)

        if resultado.error is not None:
            return resultado

        if resultado.pinch >= 1:
            return resultado

        approach += step

    return CicloOutput(fluido=resultado.fluido,
                       mezcla=resultado.mezcla,
                       water_config=water_config,
                       error="PinchBajo")

def calcular_valores_referencia(water_config: str) -> list[float]:

    # Calcular VCC de referencia
    margen_vcc = 0.3
    vcc_propano = calcular_ciclo_basico("PROPANE", [1.0],
                                        water_config).VCC
    vcc_min = (1 - margen_vcc) * vcc_propano
    vcc_max = (1 + margen_vcc) * vcc_propano
    # Calcular COP de propano
    cop_propano = calcular_ciclo_basico("PROPANE", [1.0],
                                        water_config).COP
    
    return [vcc_min, vcc_max, cop_propano]

def filtrar(resultados: list[CicloOutput], vcc_min, vcc_max) -> list[CicloOutput]:
    filtros = [
        lambda r: r.error is None,
        lambda r: vcc_min <= r.VCC <= vcc_max,
        lambda r: r.puntos.get("2").T < 130,
        lambda r: r.puntos.get("2").P < 25,
        lambda r: r.pinch > 1,
        lambda r: r.glide[0] < 10 and r.glide[1] < 10,
    ]

    for f in filtros:
        if not resultados:
            return []
        resultados = [r for r in resultados if f(r)]

    return sorted(resultados, key = lambda r: r.COP, reverse=True)

def calcular_ciclo(fluido: str | list[str], mezcla: list[float],
                   water_config: str, approach_k: float) -> CicloOutput:
    
    resultado_basico: dict[str, Any] = {}

    resultado_basico["fluido"] = fluido
    resultado_basico["mezcla"] = mezcla
    resultado_basico["water_config"] = water_config

    temperaturas_agua = WATER_CONFIG[water_config]

    [t_hw_in, t_hw_out] = temperaturas_agua["t_hw"]
    [t_cw_in, t_cw_out] = temperaturas_agua["t_cw"]

    ap_k = approach_k
    ap_0 = 3

    SH = 5
    SUB = 1

    t3 = t_hw_in + ap_k
    T_crit = rprop(fluido, "Tcrit", mezcla, T = 0, H = 0)

    try:
        if t3 > T_crit:
            raise ErrorTemperaturaTranscritica(f"Temperatura transcrítica en el punto de descarga: {t3:.1f}ºC > {T_crit:.1f}ºC")

        PK = rprop(fluido, "P", mezcla, T = t3 + SUB, Q = 0)

        # Punto 3
        P3 = TPoint(fluido, mezcla, T = t3, P = PK)
        P3.calcular("H", "D")


        # Punto 4
        P4 = TPoint(fluido, mezcla, H = P3.H, T = t_cw_out - ap_0)
        P4.calcular("P", "H", "T")
        P0 = P4.P

        # Rendimiento isentrópico
        rend_iso_h = 0.6

        # Punto 1
        t_sat_1 = rprop(fluido, "T", mezcla, P = P0, Q = 1)
        P1 = TPoint(fluido, mezcla, T = t_sat_1 + SH, P = P0)
        P1.calcular("H", "S", "V")

        # Punto 2
        h_2_s = rprop(fluido, "H", mezcla, P = PK, S = P1.S)
        h_2 = P1.H + (h_2_s - P1.H)/rend_iso_h
        P2 = TPoint(fluido, mezcla, P = PK, H = h_2)
        P2.calcular("H", "Q", "D", "T")
        if P2.Q <= 1:
            raise ErrorPuntoBifasico("El punto de descarga cae en la zona bifásica")

        # COP y VCC
        COP = (P2.H - P3.H)/(P2.H - P1.H)
        VCC = (P2.H - P1.H)/P1.V

        # Puntos saturados
        Pk_liq_sat = TPoint(fluido, mezcla, P = PK, Q = 0)
        Pk_vap_sat = TPoint(fluido, mezcla, P = PK, Q = 1)

        P0_liq_sat = TPoint(fluido, mezcla, P = P0, Q = 0)
        P0_vap_sat = TPoint(fluido, mezcla, P = P0, Q = 1)

        puntos_saturados = [Pk_liq_sat, Pk_vap_sat, P0_liq_sat, P0_vap_sat]
        # Caudales
        P_hw_in = TPoint("WATER", P = 1, T = t_hw_in)
        P_hw_out = TPoint("WATER", P = 1, T = t_hw_out)
        P_cw_in = TPoint("ETHYLENEGLYCOL", P = 1, T = t_cw_in)
        P_cw_out = TPoint("ETHYLENEGLYCOL", P = 1, T = t_cw_out)

        # Relaciones másicas
        ratio_m_GlycolHot_R = (P2.H - P3.H)/(P_hw_out.H - P_hw_in.H)
        ratio_m_GlycolCold_R = (P1.H - P4.H)/(P_cw_in.H - P_cw_out.H)

        # Relaciones volumétricas
        ratio_v_GlycolHot_R = ratio_m_GlycolHot_R * P2.D/P_hw_in.D
        ratio_v_GlycolCold_R = ratio_m_GlycolCold_R * P3.D/P_cw_in.D

        # Pinch
        h_water_pinch = P_hw_out.H - 1/ratio_m_GlycolHot_R * (P2.H - Pk_vap_sat.H)
        P_water_pinch = TPoint("WATER", P = 1, H = h_water_pinch)
        pinch = Pk_vap_sat.T - P_water_pinch.T

        # Glide
        glide_k = Pk_vap_sat.T - Pk_liq_sat.T
        glide_0 = P0_vap_sat.T - P4.T

        puntos = {
            "1": P1,
            "2": P2,
            "3": P3,
            "4": P4,
        }
        

        resultados_adicionales = {
            "COP": COP,
            "VCC": VCC,
            "puntos": puntos,
            "puntos_sat": puntos_saturados,
            "caudales_mas": [ratio_m_GlycolHot_R, ratio_m_GlycolCold_R],
            "caudales_vol": [ratio_v_GlycolHot_R, ratio_v_GlycolCold_R],
            "pinch": pinch,
            "glide": [glide_k, glide_0],
            "approach_k": ap_k,
            "error": None,
        }



        resultado = resultado_basico | resultados_adicionales

        output = CicloOutput(**resultado)

    except ErrorPuntoBifasico:
        resultado = resultado_basico | {"error": "Bifásico"}
        output = CicloOutput(**resultado)

    except ErrorTemperaturaTranscritica:
        resultado = resultado_basico | {"error": "Transcrítico"}
        output = CicloOutput(**resultado)
    
    except ZeroDivisionError:
        resultado = resultado_basico | {"error": "División 0"}
        output = CicloOutput(**resultado)

    except RuntimeError:
        resultado = resultado_basico | {"error": "REFPROP"}
        output = CicloOutput(**resultado)

    return output

def worker_calcular(args):
    # Check REFPROP handle in the refprop_utils module (initializer sets this per process)
    import refprop_utils
    if refprop_utils.RP is None:
        raise RuntimeError("REFPROP no inicializado en el worker")

    fluido, mezcla, temperaturas_agua = args
    res = calcular_ciclo_basico(fluido, mezcla, temperaturas_agua)
    return serializar(res)

# Cálculo bruto
def calcular_mezclas(posibles_refrigerantes: list[str], water_config: str):
    fichero_json = "resultados.json"
    path_json = os.path.join("resultados_ciclo_basico", water_config, "binarias", fichero_json)
    n_calcs = 41
    # Inicializar diccionario de resultados
    resultados: dict[str, dict[str, list[CicloOutput]]] = {}
    for ref_a in posibles_refrigerantes:
        resultados[ref_a] = {}
        for ref_b in posibles_refrigerantes:
            if ref_a != ref_b:
                resultados[ref_a][ref_b] = []
                [resultados[ref_a][ref_b].append(0) for _ in range(n_calcs)]

    # Calcular mezclas de refrigerantes
    print("### CÁLCULO BRUTO ###")
    n = len(posibles_refrigerantes)
    total = n * (n - 1) // 2

    with tqdm(total = total) as pbar:
        for index_a, ref_a in enumerate(posibles_refrigerantes[:-1]):

            for ref_b in posibles_refrigerantes[index_a + 1:]:

                props_a = list(np.linspace(0, 1, n_calcs))
                props_a = [float(x) for x in props_a]
                
                mezclas: list[list[float]] = [[prop_a, 1 - prop_a] for prop_a in props_a]
                lista_inputs: list[tuple[list[float], list[float], str]] = [
                    ([ref_a, ref_b], mezcla, water_config)
                    for mezcla in mezclas
                    ]

                cpu = os.cpu_count() // 2 or 1 # Usar la mitad de núcleos de la CPU
                chunksize = 2 # Está bien para la duración de la función (aprox 1s)

                with ProcessPoolExecutor(max_workers=cpu, initializer=init_refprop) as ex:
                    res = list(ex.map(worker_calcular, lista_inputs, chunksize=chunksize)) # Devuelve ya serializado
                res: list[CicloOutput] = deserializar(res)
                for index, resultado in enumerate(res):

                    resultados[ref_a][ref_b][index] = resultado
                    resultados[ref_b][ref_a][n_calcs - 1 - index] = resultado
                
                pbar.update(1)

 
    # Guardar resultados en json
    os.makedirs(os.path.dirname(path_json), exist_ok=True)
    with open(path_json, "w", encoding="utf-8") as f:
        json.dump(serializar(resultados), f, ensure_ascii=False, indent=2)

def json_a_excel(water_config: str):
    fichero_json = "resultados.json"
    path_json = os.path.join("resultados_ciclo_basico", water_config, "binarias", fichero_json)

    fichero_excel = "resultados.xlsx"
    path_excel = os.path.join("resultados_ciclo_basico", water_config, "binarias", fichero_excel)


    # Cargar json
    with open(path_json, "r", encoding="utf-8") as f:
        data: dict[str, dict[str, dict[str, dict[str, Any]]]] = json.load(f)

    # Crear ExcelWriter
    keys_filtradas = ["COP", "VCC", "error"]
    key_composicion = "mezcla"

    with pd.ExcelWriter(path_excel, engine="openpyxl") as writer:
        for hoja, columnas in data.items():
            columnas_expandidas = []
            max_filas = 0

            # Preparar datos para cada columna
            for col_name, bloques in columnas.items():
                filas = []
                for bloque in bloques:
                    bloque = bloque["__data__"]
                    for key in keys_filtradas:
                        if key in bloque:
                            if isinstance(bloque[key], float):
                                filas.append([key, round(bloque[key], 3)])
                            else:
                                filas.append([key, bloque[key]])
                    filas.append(["", ""])  # espacio entre bloques
                columnas_expandidas.append((col_name, filas))
                max_filas = max(max_filas, len(filas))

            # Rellenar con filas vacías
            for i in range(len(columnas_expandidas)):
                col_name, filas = columnas_expandidas[i]
                filas.extend([["", ""]] * (max_filas - len(filas)))
                columnas_expandidas[i] = (col_name, filas)

            # Construir DataFrame
            df_dict = {}
            # Columna extra para "Composición" (temporal, rellena con "")
            df_dict["Composición"] = [""] * max_filas
            for col_name, filas in columnas_expandidas:
                for i, subcol in enumerate(["key", "valor"]):
                    df_dict[f"{col_name}_{subcol}"] = [fila[i] for fila in filas]

            df = pd.DataFrame(df_dict)
            df.to_excel(writer, sheet_name=hoja, index=False)

    # Formato con openpyxl hoja por hoja
    wb = load_workbook(path_excel)
    for hoja in wb.sheetnames:
        ws = wb[hoja]

        # Reconstruir columnas_expandidas y bloques
        columnas = data[hoja]
        columnas_expandidas = []
        bloques_info = []  # para composición (fila inicio, fila fin, valor)
        max_filas = 0
        for col_name, bloques in columnas.items():
            filas = []
            bloque_inicio = 2  # empieza en fila 2 porque fila 1 es header
            for bloque in bloques:
                bloque = bloque["__data__"]
                bloque_filas = []
                for key in keys_filtradas:
                    if key in bloque:
                        if isinstance(bloque[key], float):
                            bloque_filas.append([key, round(bloque[key], 3)])
                        else:
                            bloque_filas.append([key, bloque[key]])
                filas.extend(bloque_filas)
                filas.append(["", ""])  # espacio entre bloques

                # Guardar info del bloque para composición
                if key_composicion in bloque:
                    # Sacar string de composición list[float] -> str
                    string_comp = ""
                    for comp in bloque[key_composicion]:
                        string_comp += f"{(comp*100):.0f}% "

                    filas_bloque = len(bloque_filas)
                    bloques_info.append((bloque_inicio, bloque_inicio + filas_bloque - 1, string_comp[:-1]))
                bloque_inicio += len(bloque_filas) + 1

            columnas_expandidas.append((col_name, filas))
            max_filas = max(max_filas, len(filas))

        # Combinar celdas del header y renombrar columnas
        col_index = 2  # columna 1 es "Composición"
        for col_name, _ in columnas_expandidas:
            start = col_index
            end = col_index + 1
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
            ws.cell(row=1, column=start).value = col_name
            col_index += 2

        # Poner título "Composición" en la primera fila, centrado
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1)
        ws.cell(row=1, column=1).value = "Composición"

        # Añadir los valores de composición combinando verticalmente cada bloque
        for start_row, end_row, valor in bloques_info:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            ws.cell(row=start_row, column=1).value = valor

        # Centrar todas las celdas
        alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = alignment

        # Ajustar ancho columnas
        for i, column_cells in enumerate(ws.columns, 1):
            ws.column_dimensions[get_column_letter(i)].width = 12

        # Inmovilizar primar fila y columna
        ws.freeze_panes = "B2"

    wb.save(path_excel)

def json_a_excel_filtrado(water_config: str) -> None:
    PASO = 0.025

    fichero_json_filtrado = "resultados_filtrados.json"
    path_json_filtrado = os.path.join("resultados_ciclo_basico", water_config, "binarias", fichero_json_filtrado)

    fichero_excel_filtrado = "resultados_filtrados.xlsx"
    path_excel_filtrado = os.path.join("resultados_ciclo_basico", water_config, "binarias", fichero_excel_filtrado)

    # Cargar json
    with open(path_json_filtrado, "r", encoding="utf-8") as f:
        data: dict[str, dict[str, list[dict[str, Any]]]] = json.load(f)

    # Eje fijo de composiciones
    composiciones = [[round(1 - i*PASO, 3), round(i*PASO, 3)] for i in range(int(1/PASO)+1)]
    col_composicion = [f"{(c[0]*100):.1f}% {(c[1]*100):.1f}%" for c in composiciones]
    indice_composicion = {tuple(c): i for i, c in enumerate(composiciones)}
    max_filas = len(composiciones)

    keys_filtradas = ["COP", "VCC", "pinch", "puntos", "glide"]
    key_composicion = "mezcla"

    with pd.ExcelWriter(path_excel_filtrado, engine="openpyxl") as writer:
        for hoja, columnas in data.items():
            # Inicializar diccionario del DataFrame con la columna Composición
            df_dict = {"Composición": col_composicion.copy()}

            # Crear columnas vacías para cada columna original
            for col_name in columnas.keys():
                df_dict[col_name] = [""] * max_filas

            # Rellenar cada columna según la mezcla
            for col_name, bloques in columnas.items():
                for bloque in bloques:
                    bloque_data = bloque["__data__"]
                    if key_composicion not in bloque_data:
                        continue
                    mezcla = [round(x, 3) for x in bloque_data[key_composicion]]
                    if tuple(mezcla) not in indice_composicion:
                        continue
                    fila_base = indice_composicion[tuple(mezcla)]

                    # Preparar texto del bloque
                    filas_texto = []
                    for key in keys_filtradas:
                        if key not in bloque_data:
                            continue
                        if key == "glide":
                            filas_texto.append(f"glide k: {bloque_data[key][0]:.1f} ºC")
                            filas_texto.append(f"glide 0: {bloque_data[key][1]:.1f} ºC")
                        elif key == "puntos":
                            puntos = deserializar(bloque_data[key])
                            filas_texto.append(f"T dis: {puntos['2'].T:.1f} ºC")
                            filas_texto.append(f"Presión k: {puntos['2'].P:.2f} bar")
                            filas_texto.append(f"Presión 0: {puntos['1'].P:.2f} bar")
                        elif key == "pinch":
                            filas_texto.append(f"T pinch: {bloque_data[key]:.1f} ºC")
                        elif key == "VCC":
                            filas_texto.append(f"VCC: {bloque_data[key]:.2f} kJ/m3")
                        elif isinstance(bloque_data[key], float):
                            filas_texto.append(f"{key}: {round(bloque_data[key], 3)}")
                        else:
                            filas_texto.append(f"{key}: {bloque_data[key]}")

                    # Unir en un solo string y poner en la fila correspondiente
                    df_dict[col_name][fila_base] = "\n".join(filas_texto)

            df = pd.DataFrame(df_dict)
            df.to_excel(writer, sheet_name=hoja, index=False)

    # Formato con openpyxl
    wb = load_workbook(path_excel_filtrado)
    for hoja in wb.sheetnames:
        ws = wb[hoja]

        # Centrar celdas y ajustar ancho
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = alignment

        for i, column_cells in enumerate(ws.columns, 1):
            ws.column_dimensions[get_column_letter(i)].width = 25

        ws.freeze_panes = "B2"

    wb.save(path_excel_filtrado)

# Cálculo fino
def refinar_mezclas(water_config: str):
    fichero_json = "resultados.json"
    path_json = os.path.join("resultados_ciclo_basico", water_config, "binarias", fichero_json)

    fichero_json_fino = "resultados_finos.json"
    path_json_fino = os.path.join("resultados_ciclo_basico", water_config, "binarias", fichero_json_fino)
    
    fichero_txt = "resultados_finos.txt"
    path_txt = os.path.join("resultados_ciclo_basico", water_config, "binarias", fichero_txt)


    with open(path_json, "r", encoding="utf-8") as f:
        data_ini: dict[str, dict[str, list[dict[str, Any]]]] = json.load(f)

    data: dict[str, dict[str, list[CicloOutput]]] = deserializar(data_ini)

    water_config = list(list(data.values())[0].values())[0][0].water_config

    posibles_refrigerantes = list(data.keys())

    # Inicializar diccionario de resultados
    resultado_fino: dict[str, dict[str, CicloOutput]] = {}
    for ref_a in posibles_refrigerantes:
        resultado_fino[ref_a] = {}


    refrigerantes_revisados: set[str] = set()

    # Calcular VCC de referencia
    margen_vcc = 0.3
    vcc_propano = calcular_ciclo_basico("PROPANE", [1.0],
                                        water_config).VCC
    vcc_min = (1 - margen_vcc) * vcc_propano
    vcc_max = (1 + margen_vcc) * vcc_propano
    # Calcular COP de propano
    cop_propano = calcular_ciclo_basico("PROPANE", [1.0],
                                        water_config).COP

    salto = 0.025

    for ref_1, sub_data_1 in data.items():

        refrigerantes_revisados.add(ref_1)

        for ref_2, lista_resultados in sub_data_1.items():
            
            if ref_2 not in refrigerantes_revisados:

                fluido = lista_resultados[0].fluido
                # Aplicar filtros
                lista = lista_resultados
                # Quitar errores
                lista = [res for res in lista if res.error is None]
                # VCC
                if lista:
                    lista = [res for res in lista if (vcc_min <= res.VCC <= vcc_max)]
                # Temperatura descarga
                if lista:
                    lista = [res for res in lista if res.puntos["2"].T < 130]
                # Pinch > 0
                if lista:
                    lista = [res for res in lista if res.pinch > 0]
                # Glide < 10ºC
                lista = [res for res in lista if res.glide[0] < 10 and res.glide[1] < 10]
                # Guardar 2 resultados con mayor COP
                lista_resultados = sorted(lista, key = lambda resultado: resultado.COP, reverse = True)[:2]

                comps = []
                salto = 0.025
                # Comprobar cuántos resultados dan COP correcto
                if len(lista_resultados) == 2: # Si hay 2 elementos coger los que tienen más COP
                    

                    # Pero si la diferencia es muy grande buscar entre los dos puntos
                    if abs(lista_resultados[0].mezcla[0] - lista_resultados[1].mezcla[0]) <= 0.10:

                        [comp_1, comp_2] = [res.mezcla for res in lista_resultados]
                        comps.append([comp_1, comp_2])
                    
                    else:
                        # Crear composiciones de +-salto alrededor de cada punto
                        comp_1 = [max(lista_resultados[0].mezcla[0] - salto, 0), 1 - max(lista_resultados[0].mezcla[0] - salto, 0)]
                        comp_2 = [min(lista_resultados[0].mezcla[0] + salto, 1), 1 - min(lista_resultados[0].mezcla[0] + salto, 1)]
                        comp_3 = [max(lista_resultados[1].mezcla[0] - salto, 0), 1 - max(lista_resultados[1].mezcla[0] - salto, 0)]
                        comp_4 = [min(lista_resultados[1].mezcla[0] + salto, 1), 1 - min(lista_resultados[1].mezcla[0] + salto, 1)]
                        
                        comps.append([comp_1, comp_2])
                        comps.append([comp_3, comp_4])

                elif len(lista_resultados) == 1:
                    # Si solo hay un elemento ir desde +- 5% composición
                    comp_1 = [max(lista_resultados[0].mezcla[0] - salto, 0), 1 - max(lista_resultados[0].mezcla[0] - salto, 0)]
                    comp_2 = [min(lista_resultados[0].mezcla[0] + salto, 1), 1 - min(lista_resultados[0].mezcla[0] + salto, 1)]
                    comps.append([comp_1, comp_2])

                else: # len(lista_resultados) == 0
                    # Si no hay resultados saltar la mezcla
                    continue
            
                # Iterar para cada rango de composiciones (solo útil cuando 
                # las 2 soluciones con más COP están alejadas y hay que 
                # crear 2 rangos diferentes)
                for comp_i in comps:
                    # Crear rango de composiciones que vaya desde comp1 a comp2
                    [comp_1, comp_2] = comp_i
                    salto = 0.005
                    range_comp: list[float] = [min(comp_1[0], comp_2[0]) + x * salto for x in range(int(round(abs(comp_1[0] - comp_2[0])/salto)) + 1)]

                    resultados: list[CicloOutput] = []

                    for comp in range_comp:
                        mezcla = [comp, 1 - comp]
                        resultado = calcular_ciclo_basico(fluido, mezcla, water_config) # BUG calculant el COP hi ha una divisió per 0 ../(PS.H - P1.H)
                        resultados.append(resultado)

                        string_comp = ""
                        if resultado.error is None:
                            for fluid, comp in zip(resultado.fluido, resultado.mezcla):
                                string_comp += f"{fluid}: {(comp*100):.1f}%, "
                            print(string_comp + f"COP = {resultado.COP:.3f}")
                        else:
                            for fluid, comp in zip(resultado.fluido, resultado.mezcla):
                                string_comp += f"{fluid}: {(comp*100):.1f}%, "
                            print(string_comp + f"ERROR = {resultado.error}")

                
                # Filtrar otra vez por errores, vcc, temp. descarga, pinch y glide
                resultados_temp_1 = [res for res in resultados if res.error is None]
                resultados_temp_2 = [res for res in resultados_temp_1 if (vcc_min <= res.VCC <= vcc_max)]
                resultados_temp_1 = [res for res in resultados_temp_2 if res.puntos["2"].T < 130]
                resultados_temp_2 = [res for res in resultados_temp_1 if res.pinch > 0]
                resultados_temp_1 = [res for res in resultados_temp_2 if res.glide[0] < 10 and res.glide[1] < 10]
                resultados_temp_2 = [res for res in resultados_temp_1 if res.puntos["2"].P < 25] 
                
                lista_mejores_resultados = []
                # Quedarse con el COP más grande
                if resultados_temp_2 != []:

                    res_mayor_COP: CicloOutput = max(resultados_temp_2, key=lambda r: r.COP)
                    
                    resultado_fino[ref_1][ref_2] = res_mayor_COP
                    resultado_fino[ref_2][ref_1] = res_mayor_COP

                    string_comp = ""
                    for fluid, comp in zip(res_mayor_COP.fluido, res_mayor_COP.mezcla):
                        string_comp += f"{fluid}: {(comp*100):.1f}%, "
                    print("\nProporción VÁLIDA con más COP: " + string_comp + f"COP = {res_mayor_COP.COP:.3f}")
                    if res_mayor_COP.COP > cop_propano:
                        print(f"COP {((res_mayor_COP.COP/cop_propano-1)*100):.1f}% más GRANDE que propano\n")
                    elif res_mayor_COP.COP < cop_propano:
                        print(f"COP {(-(res_mayor_COP.COP/cop_propano-1)*100):.1f}% más PEQUEÑO que propano\n")
                    else: 
                        print(f"COP igual al Propano: {res_mayor_COP.COP:.3f}")


    # Guardar resultados en json
    os.makedirs(os.path.dirname(path_json_fino), exist_ok=True)
    with open(path_json_fino, "w", encoding = "utf-8") as f:
        json.dump(serializar(resultado_fino), f, ensure_ascii=False, indent=2)

    # Guardar resultado en txt
    string_res: list[str] = []

    for res in lista_mejores_resultados:
        string_comp = ""

        for fluid, comp in zip(res.fluido, res.mezcla):
            string_comp += f"{fluid}: {abs((comp*100)):.0f}%, "

        string_comp = string_comp[:-2]

        proporcion = (res.COP / cop_propano - 1) * 100

        if proporcion >= 0:
            string_comp += f"\nCOP {proporcion:.2f}% más GRANDE que el propano\n\n"
        else:
            string_comp += f"\nCOP {-proporcion:.2f}% más PEQUEÑO que el propano\n\n"

        string_res.append(string_comp)

    with open(path_txt, "w",encoding="utf-8") as f:
        f.writelines(string_res)

def json_a_excel_fino(
    water_config: str,
    ancho_col_key: float = 20,
    ancho_col_value: float = 30,
    ancho_col_separador: float = 5,
) -> None:

    fichero_json_fino = "resultados_finos.json"
    path_json_fino = os.path.join("resultados_ciclo_basico", water_config, "binarias", fichero_json_fino)

    fichero_excel_fino = "resultados_finos.xlsx"
    path_excel_fino = os.path.join("resultados_ciclo_basico", water_config, "binarias", fichero_excel_fino)

    with open(path_json_fino, "r", encoding="utf-8") as f:
        data: dict[str, dict[str, dict[str, Any]]] = json.load(f)

    keys_filtradas = ["COP", "VCC", "pinch", "glide"]

    wb = Workbook()
    ws = wb.active

    align_center = Alignment(horizontal="center", vertical="center")
    bold = Font(bold=True)

    col_actual = 1

    for titulo_grupo, bloques in data.items():
        # Anchos de columnas del grupo
        ws.column_dimensions[get_column_letter(col_actual)].width = ancho_col_key
        ws.column_dimensions[get_column_letter(col_actual + 1)].width = ancho_col_value

        # Título de grupo
        ws.merge_cells(
            start_row=1,
            start_column=col_actual,
            end_row=1,
            end_column=col_actual + 1
        )
        cell = ws.cell(row=1, column=col_actual, value=titulo_grupo)
        cell.alignment = align_center
        cell.font = bold

        fila = 2

        for titulo_bloque, pares in bloques.items():
            pares = pares["__data__"]
            if pares != {}:
                # Título del bloque
                ws.merge_cells(
                    start_row=fila,
                    start_column=col_actual,
                    end_row=fila,
                    end_column=col_actual + 1
                )
                cell = ws.cell(row=fila, column=col_actual, value=titulo_bloque)
                cell.alignment = align_center
                fila += 1

                fluido = pares["fluido"]
                mezcla = pares["mezcla"]

                str_mezcla = ""
                for prop, ref in zip(mezcla, fluido):
                    str_mezcla += f"{(prop*100):.1f}% {ref} + "
                str_mezcla = str_mezcla[:-3]

                # Contenido
                for k in pares:
                    if k not in keys_filtradas:
                        continue
                    c_key = ws.cell(row=fila, column=col_actual, value=k)
                    if k == "glide":
                        texto = ""
                        for v in pares[k]:
                            texto += f"{v:.2f}ºC "
                        c_val = ws.cell(row=fila, column=col_actual + 1, value= texto)
                    elif k == "pinch":
                        c_val = ws.cell(row=fila, column=col_actual + 1, value= f"{pares[k]:.2f} ºC")
                    else:
                        c_val = ws.cell(row=fila, column=col_actual + 1, value= round(pares[k], 3))
                    c_key.alignment = align_center
                    c_val.alignment = align_center
                    fila += 1
                
                c_key = ws.cell(row=fila, column=col_actual, value = "Composición")
                c_val = ws.cell(row=fila, column=col_actual + 1, value = str_mezcla)
                c_key.alignment = align_center
                c_val.alignment = align_center
                fila += 1

                fila += 1  # línea en blanco entre bloques

        # Columna separadora
        sep_col = col_actual + 2
        ws.column_dimensions[get_column_letter(sep_col)].width = ancho_col_separador

        col_actual += 3

    ws.freeze_panes = "A2"

    wb.save(path_excel_fino)

# Generar gráficos
def generar_graficos_binarios(casos, valor_referencia, water_config):
    output_folder = os.path.join("resultados_ciclo_basico", water_config, "binarias", "graficos")
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    for caso in casos:
        nombre = caso["nombre"]
        ejes = caso["ejes"]
        datos = caso["valores"]

        # 1. Extraer y ordenar datos
        sorted_keys = sorted(datos.keys())
        x_values = [k[0] for k in sorted_keys]
        
        # 2. Calcular la desviación porcentual respecto a la referencia
        # Formula: ((valor - ref) / ref) * 100
        y_percent = [((datos[k] - valor_referencia) / valor_referencia) * 100 for k in sorted_keys]

        plt.figure(figsize=(10, 6))
        
        # 3. Dibujar puntos SIN unir (scatter)
        plt.scatter(x_values, y_percent, color='b', label=f'Desviación {nombre}')
        
        # 4. Línea de referencia en el 0%
        plt.axhline(y=0, color='r', linestyle='--', label=f'Referencia ({valor_referencia:.3f})')

        # 5. Configurar ticks de los ejes
        # Eje X en porcentajes
        plt.xticks([i/10 for i in range(11)], [f'{i*10}%' for i in range(11)])
        
        # Eje Y con sufijo %
        current_values = plt.gca().get_yticks()
        plt.gca().set_yticklabels([f'{int(x)}%' for x in current_values])

        plt.title(f'Desviación porcentual respecto a referencia: {nombre}', fontsize=12)
        plt.xlabel(f'Composición de {ejes[0]}')
        plt.ylabel('Diferencia de COP respecto a referencia (%)')
        plt.grid(True, linestyle=':', alpha=0.6)
        plt.legend()
        plt.xlim(0, 1)

        plt.savefig(os.path.join(output_folder, f"{nombre}.png"))
        plt.close()

def ciclo_basico_filtrado(water_config: str) -> None:
    fichero_json = "resultados.json"
    path_json = os.path.join("resultados_ciclo_basico", water_config, "binarias", fichero_json)

    fichero_json_filtrado = "resultados_filtrados.json"
    path_json_filtrado = os.path.join("resultados_ciclo_basico", water_config, "binarias", fichero_json_filtrado)

    with open(path_json, "r", encoding="utf-8") as f:
        dic_res = json.load(f)

    dic_res: dict[str, dict[str, list[CicloOutput]]] = deserializar(dic_res)

    water_config = list(list(dic_res.values())[0].values())[0][0].water_config
    [vcc_min, vcc_max, cop_propano] = calcular_valores_referencia(water_config)

    dic_temp: dict[str, dict[str, list[CicloOutput]]] = {}
    for ref_a, sub_dict in dic_res.items():
        for ref_b, lista_res in sub_dict.items():
            dic_temp.setdefault(ref_a, {})[ref_b] = filtrar(lista_res, vcc_min, vcc_max)

    dic_res = serializar(dic_temp)

    with open(path_json_filtrado, "w", encoding="utf-8") as f:
        json.dump(dic_res, f, ensure_ascii=False, indent=4)

def crear_casos(water_config: str) -> tuple[list[dict[str, Any]], float]:
    
    fichero_json_filtrado = "resultados_filtrados.json"
    path_json_filtrado = os.path.join("resultados_ciclo_basico", water_config, "binarias", fichero_json_filtrado)

    with open(path_json_filtrado, "r", encoding="utf-8") as f:
        dic_res = json.load(f)

    dic_res: dict[str, dict[str, list[CicloOutput]]] = deserializar(dic_res)

    posibles_refrigerantes = list(dic_res.keys())

    water_config = list(list(dic_res.values())[0].values())[0][0].water_config

    [vcc_min, vcc_max, cop_propano] = calcular_valores_referencia(water_config)

    casos: list[dict[str, Any]] = []

    for index_a, ref_a in enumerate(posibles_refrigerantes[:-1]):
        for ref_b in posibles_refrigerantes[index_a + 1:]:
            if dic_res.get(ref_a, {}).get(ref_b):
                lista_resultados = dic_res[ref_a][ref_b]
                fluidos = lista_resultados[0].fluido
                ejes = fluidos
                nombre = (", ").join(fluidos)
                valores = {
                    tuple(res.mezcla): res.COP
                    for res in lista_resultados
                }
                casos.append({
                    "nombre": nombre,
                    "ejes": ejes,
                    "valores": valores,
                })

    return (casos, cop_propano)

init_refprop()
water_config = "media"

# Resumen
def crear_datos_resumen(water_config: str) -> list[dict[str, Any]]:
    fichero_json = "resultados.json"
    path_json = os.path.join("resultados_ciclo_basico", water_config, "binarias", fichero_json)

    with open(path_json, "r", encoding="utf-8") as f:
        dic_res = json.load(f)

    dic_res: dict[str, dict[str, list[CicloOutput]]] = deserializar(dic_res)

    [vcc_min, vcc_max, cop_propano] = calcular_valores_referencia(water_config)

    # Quitar valores que tengan COP menor a COP_propano
    dic_temp: dict[str, dict[str, list[CicloOutput]]] = {}
    for ref_a, sub_dict in dic_res.items():
        for ref_b, lista_res in sub_dict.items():
            dic_temp.setdefault(ref_a, {})[ref_b] = filtrar(lista_res, vcc_min, vcc_max)
            if dic_temp[ref_a][ref_b]:
                dic_temp[ref_a][ref_b] = [
                    res for res in dic_temp[ref_a][ref_b] if res.COP >= cop_propano
                ]

    dic_res = dic_temp
    del dic_temp

    posibles_refrigerantes = list(dic_res.keys())

    lista_mejores_res: list[list[CicloOutput]] = []

    for index_a, ref_a in enumerate(posibles_refrigerantes[:-1]):

        for ref_b in posibles_refrigerantes[index_a + 1:]:

            lista_mejores_res.append(dic_res.get(ref_a, {}).get(ref_b, [])) if dic_res.get(ref_a, {}).get(ref_b, []) else ...

    # Quitar listas vacías
    lista_mejores_res = [sub for sub in lista_mejores_res if sub]

    # Ordenar por COP
    lista_mejores_res.sort(key = lambda list_res: max(res.COP for res in list_res), reverse=True)

    datos_resumen: list[dict[str, Any]] = []

    def resumen(lista: list[float]) -> list[float]:
        return [lista[0], lista[-1], max(lista), min(lista)]

    for list_res in lista_mejores_res:
        [pprint(res.mezcla) for res in list_res]
        print()
        # BUG
        fluido = lista_res[0].fluido
        mezcla = [
            [round(v, 3) for v in res.mezcla]
            for res in [lista_res[0], lista_res[-1]]
        ]
        mezcla = [f"{mezcla[0][0]*100:.1f}% {fluido[0]} + {mezcla[0][1]*100:.1f}% {fluido[1]}", f"{mezcla[1][0]*100:.1f}% {fluido[0]} + {mezcla[1][1]*100:.1f}% {fluido[1]}"]

        

        cop = resumen([res.COP for res in list_res])
        vcc = resumen([res.VCC for res in list_res])
        pinch = resumen([res.pinch for res in list_res])
        T_dis = resumen([res.puntos["2"].T for res in list_res])
        p_k = resumen([res.puntos["2"].P for res in list_res])
        p_0 = resumen([res.puntos["1"].P for res in list_res])
        glide_k = resumen([res.glide[0] for res in list_res])
        glide_0 = resumen([res.glide[1] for res in list_res])

        datos_resumen.append({
            "fluido": ", ".join(list_res[0].fluido),
            "mezcla": mezcla,
            "COP": cop,
            "VCC": vcc,
            "T pinch": pinch,
            "T descarga": T_dis,
            "Presion k": p_k,
            "Presion 0": p_0,
            "glide k": glide_k,
            "glide 0": glide_0,
        })
    
    return datos_resumen

def crear_excel(
    datos: list[dict[str, Any]],
    ancho_columna: float = 14,
    ancho_separador: float = 4
):
    fichero_excel_resumen = "resumen_resultados.xlsx"
    path_excel_resumen = os.path.join(
        "resultados_ciclo_basico",
        water_config,
        "binarias",
        fichero_excel_resumen
    )

    FORMATO_KEYS = {
    "Presion 0":     {"dec": 2, "unit": " bar"},
    "Presion k":     {"dec": 2, "unit": " bar"},
    "T descarga":   {"dec": 1, "unit": " °C"},
    "COP":     {"dec": 2, "unit": ""},
    "VCC":     {"dec": 1, "unit": " kJ/m3"},
    "T pinch":   {"dec": 2, "unit": " ºC"},
    "glide 0": {"dec": 2, "unit": " ºC"},
    "glide k": {"dec": 2, "unit": " ºC"},
}

    def formatear_valor(key: str, valor: float) -> str:
        cfg = FORMATO_KEYS.get(key)
        if cfg is None:
            return valor

        dec = cfg["dec"]
        unit = cfg.get("unit", "")

        if unit:
            return f"{valor:.{dec}f}{unit}"
        else:
            return round(valor, dec)

    def autoajustar_columnas(ws, col_inicio, col_relativas, fila_inicio=1, fila_fin=None, padding=2):
        if fila_fin is None:
            fila_fin = ws.max_row

        for offset in col_relativas:
            col_idx = col_inicio + offset
            col_letter = get_column_letter(col_idx)

            max_len = 0
            for fila in range(fila_inicio, fila_fin + 1):
                value = ws.cell(row=fila, column=col_idx).value
                if value is not None:
                    max_len = max(max_len, len(str(value)))

            ws.column_dimensions[col_letter].width = max_len + padding

    os.makedirs(os.path.dirname(path_excel_resumen), exist_ok=True)

    wb = Workbook()
    ws = wb.active

    align_center = Alignment(horizontal="center", vertical="center")
    col_inicio = 1

    for bloque in datos:
        # Anchos columnas bloque
        for i in range(5):
            col = get_column_letter(col_inicio + i)
            ws.column_dimensions[col].width = ancho_columna

        # Fila 1: fluido
        ws.merge_cells(
            start_row=1,
            start_column=col_inicio,
            end_row=1,
            end_column=col_inicio + 4
        )
        ws.cell(row=1, column=col_inicio, value=bloque["fluido"]).alignment = align_center

        # Fila 2
        headers = ["", "Inicial", "Final", "Máximo", "Mínimo"]
        for i, h in enumerate(headers):
            ws.cell(row=2, column=col_inicio + i, value=h).alignment = align_center

        # Fila 3: mezcla
        ws.cell(row=3, column=col_inicio, value="mezcla").alignment = align_center
        ws.cell(row=3, column=col_inicio + 1, value=bloque["mezcla"][0]).alignment = align_center
        ws.cell(row=3, column=col_inicio + 2, value=bloque["mezcla"][1]).alignment = align_center

        fila = 4
        for k, v in bloque.items():
            if isinstance(v, list) and len(v) == 4:
                ws.cell(row=fila, column=col_inicio, value=k).alignment = align_center
                for i in range(4):
                    valor = formatear_valor(k, v[i])
                    ws.cell(
                        row=fila,
                        column=col_inicio + 1 + i,
                        value=valor
                    ).alignment = align_center
                fila += 1

        # Columna separadora
        sep_col = get_column_letter(col_inicio + 5)
        ws.column_dimensions[sep_col].width = ancho_separador

        autoajustar_columnas(
            ws,
            col_inicio=col_inicio,
            col_relativas=[1, 2]  # columnas 2 y 3 del bloque
        )

        col_inicio += 6
        
    wb.save(path_excel_resumen)
    wb.close()

datos_resumen = crear_datos_resumen(water_config)

crear_excel(datos_resumen)

def main():
    init_refprop()
    
    # DATOS
    water_config = "baja" # "baja" / "intermedia" / "media" / "alta"

    posibles_refrigerantes = ["PROPANE", "BUTANE", "ISOBUTANE", "PROPYLENE", "DME"]


    # CÁLCULO BRUTO
    calcular_mezclas(posibles_refrigerantes, water_config)

    json_a_excel(water_config)

    # CÁLCULO FINO
    refinar_mezclas(water_config)

    json_a_excel_fino(water_config)

    # CREAR GRÁFICOS BINARIOS
    ciclo_basico_filtrado(water_config)

    json_a_excel_filtrado(water_config)

    (casos, cop_propano) = crear_casos(water_config)

    generar_graficos_binarios(casos, cop_propano, water_config)

    # CREAR RESUMEN
    datos_resumen = crear_datos_resumen(water_config)

    crear_excel(datos_resumen)



if __name__ == "__main__":
    ...


