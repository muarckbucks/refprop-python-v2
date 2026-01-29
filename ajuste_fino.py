import json
from typing import Any, Iterable
from refprop_utils import *
from ciclo_basico import calcular_ciclo_basico
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter




def refinar_mezclas(fichero_json, fichero_json_fino):
    with open(fichero_json, "r", encoding="utf-8") as f:
        data: dict[str, dict[str, list[dict[str, Any]]]] = json.load(f)

    posibles_refrigerantes = list(data.keys())

    # Inicializar diccionario de resultados
    resultado_fino: dict[str, dict[str, dict[str, Any]]] = {}
    for index_a, ref_a in enumerate(posibles_refrigerantes):
        resultado_fino[ref_a] = {}
        for index_b, ref_b in enumerate(posibles_refrigerantes):
            if index_a != index_b:
                resultado_fino[ref_a][ref_b] = {}

    refrigerantes_revisados: set[str] = set()

    for index_1, [ref_1, sub_data_1] in enumerate(data.items()):

        refrigerantes_revisados.add(ref_1)

        for ref_2, lista_resultados in sub_data_1.items():
            
            if ref_2 not in refrigerantes_revisados:

                fluido = lista_resultados[0]["fluido"]
                temperaturas_agua = lista_resultados[0]["temperaturas agua"]
                # Guardar 2 resultados con mayor COP
                lista_res_temp = [res for res in lista_resultados if isinstance(res["COP"], float)]
                lista_resultados = sorted(lista_res_temp, key = lambda resultado: resultado["COP"], reverse = True)[:2]

                comps = []

                # Comprobar cuántos resultados dan COP correcto
                if len(lista_resultados) == 2: # Si hay 2 elementos coger los que tienen más COP
                    

                    # Pero si la diferencia es muy grande buscar entre los dos puntos
                    if abs(lista_resultados[0]["mezcla"][0] - lista_resultados[1]["mezcla"][0]) <= 0.10:

                        [comp_1, comp_2] = [res["mezcla"] for res in lista_resultados]
                        comps.append([comp_1, comp_2])
                    
                    else:
                        # Crear composiciones de +-5% alrededor de cada punto
                        comp_1 = [max(lista_resultados[0]["mezcla"][0] - 0.05, 0), 1 - max(lista_resultados[0]["mezcla"][0] - 0.05, 0)]
                        comp_2 = [min(lista_resultados[0]["mezcla"][0] + 0.05, 1), 1 - min(lista_resultados[0]["mezcla"][0] + 0.05, 1)]
                        comp_3 = [max(lista_resultados[1]["mezcla"][0] - 0.05, 0), 1 - max(lista_resultados[1]["mezcla"][0] - 0.05, 0)]
                        comp_4 = [min(lista_resultados[1]["mezcla"][0] + 0.05, 1), 1 - min(lista_resultados[1]["mezcla"][0] + 0.05, 1)]
                        
                        comps.append([comp_1, comp_2])
                        comps.append([comp_3, comp_4])

                elif len(lista_resultados) == 1:
                    # Si solo hay un elemento ir desde +- 5% composición
                    comp_1 = [min(lista_resultados[0]["mezcla"][0] - 0.05, 0), 1 - min(lista_resultados[0]["mezcla"][0] - 0.05, 0)]
                    comp_2 = [max(lista_resultados[0]["mezcla"][0] + 0.05, 1), 1 - max(lista_resultados[0]["mezcla"][0] + 0.05, 1)]
                    comps.append([comp_1, comp_2])

                else: # len(lista_resultados) == 0
                    # Si no hay resultados saltar la mezcla
                    continue
            
                # Iterar para cada rango de composiciones (solo útil cuando 
                # las 2 soluciones con más COP están alejadas y hay que 
                # crear 2 rangos diferentes)

                for comp_i in comps:
                    # Crear rango de composiciones que vaya desde comp1 a comp2
                    [comp_1, comp_2] = comp_i
                    salto = 0.005
                    range_comp: list[float] = [min(comp_1[0], comp_2[0]) + x * salto for x in range(int(round(abs(comp_1[0] - comp_2[0])/salto)) + 1)]

                    resultados: list[dict, Any] = []

                    for comp in range_comp:
                        mezcla = [comp, 1 - comp]
                        resultado = calcular_ciclo_basico(fluido, mezcla, temperaturas_agua) # BUG calculant el COP hi ha una divisió per 0 ../(PS.H - P1.H)
                        resultados.append(resultado)

                        string_comp = ""
                        if resultado["error"] != "-":
                            for fluid, comp in zip(resultado["fluido"], resultado["mezcla"]):
                                string_comp += f"{fluid}: {(comp*100):.1f}%, "
                            print(string_comp + f"COP = {resultado["COP"]:.3f}")
                        else:
                            for fluid, comp in zip(resultado["fluido"], resultado["mezcla"]):
                                string_comp += f"{fluid}: {(comp*100):.1f}%, "
                            print(string_comp + f"ERROR = {resultado["error"]}")

                # Quedarse con el COP más grande

                resultados = [res for res in resultados if res["error"] == "-"]
                res_mayor_COP = sorted(resultados, key = lambda resultado: resultado["COP"], reverse = True)[0]

                resultado_fino[ref_1][ref_2] |= res_mayor_COP
                resultado_fino[ref_2][ref_1] |= res_mayor_COP

                string_comp = ""
                for fluid, comp in zip(res_mayor_COP["fluido"], res_mayor_COP["mezcla"]):
                    string_comp += f"{fluid}: {(comp*100):.1f}%, "
                print("\nProporción con más COP: " + string_comp + f"COP = {res_mayor_COP["COP"]:.3f}\n")


    # Guardar resultados en json
    claves_permitidas = {"fluido", "mezcla", "presiones", "caudales másicos", "caudales volumétricos", "COP", "VCC", "pinch", "glide", "error", "temperaturas agua"}
    def filtrar_resultados(
        resultados: dict[str, dict[str, dict[str, Any]]],
        claves_permitidas: set[str]
    ) -> dict[str, dict[str, dict[str, Any]]]:
        return {
            k1: {
                k2: {
                    kk: vv for kk, vv in d.items()
                    if kk in claves_permitidas
                }
                for k2, d in sub.items()
            }
            for k1, sub in resultados.items()
        }

    filtrados = filtrar_resultados(resultado_fino, claves_permitidas)

    with open(fichero_json_fino, "w", encoding = "utf-8") as f:
        json.dump(filtrados, f, ensure_ascii=False, indent=2)

def json_a_excel(
    fichero_json: str,
    fichero_excel: str,
    keys_filtradas: Iterable[str],
    ancho_col_key: float = 20,
    ancho_col_value: float = 30,
    ancho_col_separador: float = 5,
) -> None:
    """
    - Cada grupo ocupa 2 columnas (key / value)
    - Columna en blanco entre grupos con ancho configurable
    - Bloques en vertical, separados por una fila en blanco
    - Primera fila (títulos de grupo) en negrita
    """

    keys_filtradas = list(keys_filtradas)

    with open(fichero_json, "r", encoding="utf-8") as f:
        data: dict[str, dict[str, dict[str, Any]]] = json.load(f)

    wb = Workbook()
    ws = wb.active

    align_center = Alignment(horizontal="center", vertical="center")
    bold = Font(bold=True)

    col_actual = 1

    for titulo_grupo, bloques in data.items():
        # Anchos de columnas del grupo
        ws.column_dimensions[get_column_letter(col_actual)].width = ancho_col_key
        ws.column_dimensions[get_column_letter(col_actual + 1)].width = ancho_col_value

        # Título de grupo
        ws.merge_cells(
            start_row=1,
            start_column=col_actual,
            end_row=1,
            end_column=col_actual + 1
        )
        cell = ws.cell(row=1, column=col_actual, value=titulo_grupo)
        cell.alignment = align_center
        cell.font = bold

        fila = 2

        for titulo_bloque, pares in bloques.items():
            if pares != {}:
                # Título del bloque
                ws.merge_cells(
                    start_row=fila,
                    start_column=col_actual,
                    end_row=fila,
                    end_column=col_actual + 1
                )
                cell = ws.cell(row=fila, column=col_actual, value=titulo_bloque)
                cell.alignment = align_center
                fila += 1

                fluido = pares["fluido"]
                mezcla = pares["mezcla"]

                str_mezcla = ""
                for prop, ref in zip(mezcla, fluido):
                    str_mezcla += f"{(prop*100):.1f}% {ref} + "
                str_mezcla = str_mezcla[:-3]

                # Contenido
                for k in pares:
                    if k not in keys_filtradas:
                        continue
                    c_key = ws.cell(row=fila, column=col_actual, value=k)
                    c_val = ws.cell(row=fila, column=col_actual + 1, value= round(pares[k], 3))
                    c_key.alignment = align_center
                    c_val.alignment = align_center
                    fila += 1
                
                c_key = ws.cell(row=fila, column=col_actual, value = "Composición")
                c_val = ws.cell(row=fila, column=col_actual + 1, value = str_mezcla)
                c_key.alignment = align_center
                c_val.alignment = align_center
                fila += 1

                fila += 1  # línea en blanco entre bloques

        # Columna separadora
        sep_col = col_actual + 2
        ws.column_dimensions[get_column_letter(sep_col)].width = ancho_col_separador

        col_actual += 3

    ws.freeze_panes = "A2"

    wb.save(fichero_excel)


def main():

    fichero_json = r"resultados\resultados_ciclo_basico.json"
    fichero_json_fino = r"resultados\resultados_finos_basico.json"
    fichero_excel = r"resultados\resultados_finos_basico.xlsx"

    keys_permitidas = ["COP", "VCC"]

    ancho_col_key = 20
    ancho_col_value = 30
    ancho_col_separador = 5

    refinar_mezclas(fichero_json, fichero_json_fino)
    json_a_excel(fichero_json_fino, fichero_excel, keys_permitidas,
                 ancho_col_key, ancho_col_value, ancho_col_separador)


if __name__ == "__main__":
    main()
            

            