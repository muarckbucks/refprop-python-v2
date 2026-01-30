from refprop_utils import *
from ciclo_basico import *
import numpy as np
import json
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from typing import Any

Cliente = ClienteRefprop(r"C:\Program Files (x86)\REFPROP\REFPRP64.DLL")

def calcular_mezclas(fichero_json, posibles_refrigerantes, temperaturas_agua):

    # Inicializar diccionario de resultados
    resultados: dict[str, dict[str, list[dict[str, Any]]]] = {}
    for index_a, ref_a in enumerate(posibles_refrigerantes):
        resultados[ref_a] = {}
        for index_b, ref_b in enumerate(posibles_refrigerantes):
            if index_a != index_b:
                resultados[ref_a][ref_b] = []
                [resultados[ref_a][ref_b].append({}) for _ in range(n_calcs)]

    # Calcular mezclas de refrigerantes
    for index_a, ref_a in enumerate(posibles_refrigerantes[:-1]):

        for index_b, ref_b in enumerate(posibles_refrigerantes[index_a+1:]):

            props_a = list(np.linspace(0, 1, n_calcs))
            props_a = [float(x) for x in props_a]
            
            for index_prop, prop_a in enumerate(props_a):
                prop_b = 1 - prop_a
                mezcla = [prop_a, prop_b]


                resultado = calcular_ciclo_basico([ref_a, ref_b], mezcla, temperaturas_agua)
                # Ir imprimiendo resultados
                string_comp = ""

                # Comprobar si no ha dado error el cálculo
                if "glide" in resultado:
                    for fluid, comp in zip(resultado["fluido"], resultado["mezcla"]):
                        string_comp += f"{fluid}: {(comp*100):.0f}%, "
                    print(string_comp + f"COP = {resultado["COP"]:.3f}")
                else:
                    for fluid, comp in zip(resultado["fluido"], resultado["mezcla"]):
                        string_comp += f"{fluid}: {(comp*100):.0f}%, "
                    print(string_comp + f"ERROR = {resultado["error"]}")                    
                # puntos_PH(resultado["puntos"], 1.5, 0.2)

                resultados[ref_a][ref_b][index_prop] |= resultado
                resultados[ref_b][ref_a][n_calcs - 1 - index_prop] |= resultado

        
    # Guardar resultados en json
    claves_permitidas = {"fluido", "mezcla", "presiones", "entalpias", "temperaturas", "caudales másicos", "caudales volumétricos", "COP", "VCC", "pinch", "glide", "error", "temperaturas agua"}
    def filtrar_resultados(resultados: dict[str, dict[str, list[dict[str, Any]]]],
                           claves_permitidas = set[str]) -> dict[str, dict[str, list[dict[str, Any]]]]:
        return {
            k1 : {
                k2 : [
                    {kk: vv for kk, vv in d.items() if kk in claves_permitidas}
                    for d in lista
                ]
                for k2, lista in sub.items()

            }
            for k1, sub in resultados.items()
        }

    filtrados = filtrar_resultados(resultados, claves_permitidas)

    with open(fichero_json, "w", encoding="utf-8") as f:
        json.dump(filtrados, f, ensure_ascii=False, indent=2)

def json_a_excel(fichero_json, fichero_excel):

    # Cargar json
    with open(fichero_json, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Crear ExcelWriter
    keys_filtradas = ["COP", "VCC", "error"]
    key_composicion = "mezcla"

    with pd.ExcelWriter(fichero_excel, engine="openpyxl") as writer:
        for hoja, columnas in data.items():
            columnas_expandidas = []
            max_filas = 0

            # Preparar datos para cada columna
            for col_name, bloques in columnas.items():
                filas = []
                for bloque in bloques:
                    for key in keys_filtradas:
                        if key in bloque:
                            if isinstance(bloque[key], float):
                                filas.append([key, round(bloque[key], 3)])
                            else:
                                filas.append([key, bloque[key]])
                    filas.append(["", ""])  # espacio entre bloques
                columnas_expandidas.append((col_name, filas))
                max_filas = max(max_filas, len(filas))

            # Rellenar con filas vacías
            for i in range(len(columnas_expandidas)):
                col_name, filas = columnas_expandidas[i]
                filas.extend([["", ""]] * (max_filas - len(filas)))
                columnas_expandidas[i] = (col_name, filas)

            # Construir DataFrame
            df_dict = {}
            # Columna extra para "Composición" (temporal, rellena con "")
            df_dict["Composición"] = [""] * max_filas
            for col_name, filas in columnas_expandidas:
                for i, subcol in enumerate(["key", "valor"]):
                    df_dict[f"{col_name}_{subcol}"] = [fila[i] for fila in filas]

            df = pd.DataFrame(df_dict)
            df.to_excel(writer, sheet_name=hoja, index=False)

    # Formato con openpyxl hoja por hoja
    wb = load_workbook(fichero_excel)
    for hoja in wb.sheetnames:
        ws = wb[hoja]

        # Reconstruir columnas_expandidas y bloques
        columnas = data[hoja]
        columnas_expandidas = []
        bloques_info = []  # para composición (fila inicio, fila fin, valor)
        max_filas = 0
        for col_name, bloques in columnas.items():
            filas = []
            bloque_inicio = 2  # empieza en fila 2 porque fila 1 es header
            for bloque in bloques:
                bloque_filas = []
                for key in keys_filtradas:
                    if key in bloque:
                        if isinstance(bloque[key], float):
                            bloque_filas.append([key, round(bloque[key], 3)])
                        else:
                            bloque_filas.append([key, bloque[key]])
                filas.extend(bloque_filas)
                filas.append(["", ""])  # espacio entre bloques

                # Guardar info del bloque para composición
                if key_composicion in bloque:
                    # Sacar string de composición list[float] -> str
                    string_comp = ""
                    for comp in bloque[key_composicion]:
                        string_comp += f"{(comp*100):.0f}% "

                    filas_bloque = len(bloque_filas)
                    bloques_info.append((bloque_inicio, bloque_inicio + filas_bloque - 1, string_comp[:-1]))
                bloque_inicio += len(bloque_filas) + 1

            columnas_expandidas.append((col_name, filas))
            max_filas = max(max_filas, len(filas))

        # Combinar celdas del header y renombrar columnas
        col_index = 2  # columna 1 es "Composición"
        for col_name, _ in columnas_expandidas:
            start = col_index
            end = col_index + 1
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
            ws.cell(row=1, column=start).value = col_name
            col_index += 2

        # Poner título "Composición" en la primera fila, centrado
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1)
        ws.cell(row=1, column=1).value = "Composición"

        # Añadir los valores de composición combinando verticalmente cada bloque
        for start_row, end_row, valor in bloques_info:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            ws.cell(row=start_row, column=1).value = valor

        # Centrar todas las celdas
        alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = alignment

        # Ajustar ancho columnas
        for i, column_cells in enumerate(ws.columns, 1):
            ws.column_dimensions[get_column_letter(i)].width = 12

        # Inmovilizar primar fila y columna
        ws.freeze_panes = "B2"

    wb.save(fichero_excel)

def main(fichero_json, fichero_excel, posibles_refrigerantes, temperaturas_agua):
    calcular_mezclas(fichero_json, posibles_refrigerantes, temperaturas_agua)
    json_a_excel(fichero_json, fichero_excel)


if __name__ == "__main__":
    
    # DATOS BÁSICOS
    t_hw_in = 47
    t_hw_out = 55

    t_cw_in = 0
    t_cw_out = -3

    temperaturas_agua = {
        "t_hw": [t_hw_in, t_hw_out],
        "t_cw": [t_cw_in, t_cw_out]
    }

    n_calcs = 21 # Para que haya diferencia de 5%

    posibles_refrigerantes = ["PROPANE", "CO2", "BUTANE", "ISOBUTANE", "PROPYLENE",
                              "PENTANE", "DME", "ETHANE", "HEXANE", "TOLUENE"]

    fichero_json = r"resultados\resultados_ciclo_basico.json"
    fichero_excel = r"resultados\resultados_ciclo_basico.xlsx"
    
    
    main(fichero_json, fichero_excel, posibles_refrigerantes, temperaturas_agua)
    
